"""openpyxlを使ってXLSXを安全に読み取るアダプター。"""

import importlib.util
import re
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass
from importlib import metadata
from pathlib import Path
from typing import Any, Protocol, cast

from signate_drive_rag.domain import JsonValue

OPENPYXL_LICENSE = "MIT"
DEFUSEDXML_LICENSE = "PSF"


@dataclass(frozen=True, slots=True)
class XlsxParserOptions:
    """XLSX抽出時の安全上限と分割条件。"""

    max_uncompressed_bytes: int = 2_000_000_000
    max_compression_ratio: float = 1000.0
    metadata_inspection_max_file_bytes: int = 20_000_000
    max_rows_per_unit: int = 25
    max_columns_per_unit: int = 20
    max_characters_per_unit: int = 4_000
    max_cells_per_unit: int = 500
    large_sheet_row_threshold: int = 10_000
    very_wide_sheet_column_threshold: int = 100
    large_cell_value_characters: int = 4_000


@dataclass(frozen=True, slots=True)
class DefinedNameInfo:
    """Excel定義名の追跡用情報。"""

    defined_name: str
    destination_sheet: str | None
    destination_range: str | None


@dataclass(frozen=True, slots=True)
class XlsxTableInfo:
    """Excelテーブルの追跡用情報。"""

    table_name: str
    display_name: str
    range: str
    header_row: int | None


@dataclass(frozen=True, slots=True)
class SheetInspection:
    """ワークシートの抽出前診断結果。"""

    sheet_name: str
    sheet_index: int
    sheet_state: str
    declared_dimension: str
    actual_dimension: str | None
    min_row: int | None
    max_row: int | None
    min_column: int | None
    max_column: int | None
    non_empty_cell_count: int
    formula_cell_count: int
    merged_ranges: tuple[str, ...]
    tables: tuple[XlsxTableInfo, ...]
    hidden_rows: tuple[int, ...]
    hidden_columns: tuple[int, ...]


@dataclass(frozen=True, slots=True)
class WorkbookInspection:
    """XLSXワークブック全体の診断結果。"""

    sheets: tuple[SheetInspection, ...]
    defined_names: tuple[DefinedNameInfo, ...]
    formulas: dict[str, dict[str, str]]
    metadata_limited: bool
    openpyxl_version: str
    defusedxml_available: bool
    openpyxl_license: str
    defusedxml_license: str
    zip_uncompressed_bytes: int
    zip_compressed_bytes: int
    zip_compression_ratio: float


@dataclass(frozen=True, slots=True)
class SpreadsheetCell:
    """ストリーミング取得したセル値。"""

    row_number: int
    column_number: int
    coordinate: str
    value: object


@dataclass(frozen=True, slots=True)
class SpreadsheetRow:
    """ストリーミング取得した1行分のセル値。"""

    sheet_name: str
    row_number: int
    cells: tuple[SpreadsheetCell, ...]


class XlsxWorkbookAdapter(Protocol):
    """XLSXワークブックをストリーミング形式で読み取る。"""

    def inspect(self, source_path: Path, options: XlsxParserOptions) -> WorkbookInspection:
        """XLSXコンテナとワークブック構造を診断する。"""
        ...

    def iter_sheet_rows(
        self,
        source_path: Path,
        sheet_name: str,
        inspection: SheetInspection,
    ) -> Iterator[SpreadsheetRow]:
        """指定シートの使用範囲を行単位で読み取る。"""
        ...


class XlsxWorkbookError(RuntimeError):
    """XLSXワークブックを安全に読み取れない場合の例外。"""

    def __init__(self, issue_type: str, message: str) -> None:
        """抽出失敗へ変換しやすいissue_type付き例外を作成する。"""
        super().__init__(message)
        self.issue_type = issue_type


class OpenpyxlWorkbookAdapter:
    """openpyxl固有型を抽出パーサー外へ広げないためのアダプター。"""

    def inspect(self, source_path: Path, options: XlsxParserOptions) -> WorkbookInspection:
        """OOXML診断後に、数式と軽量metadataを取得する。"""
        zip_statistics = _inspect_ooxml_container(source_path, options)
        formulas, formula_sheet_inspections = _collect_formulas_and_sheet_metrics(source_path)
        metadata_limited = source_path.stat().st_size > options.metadata_inspection_max_file_bytes
        metadata_by_sheet: dict[str, _SheetMetadata] = {}
        defined_names: tuple[DefinedNameInfo, ...] = ()
        if not metadata_limited:
            metadata_by_sheet, defined_names = _collect_workbook_metadata(source_path)

        sheets = tuple(
            _merge_sheet_inspection(
                sheet,
                metadata_by_sheet.get(sheet.sheet_name),
            )
            for sheet in formula_sheet_inspections
        )
        return WorkbookInspection(
            sheets=sheets,
            defined_names=defined_names,
            formulas=formulas,
            metadata_limited=metadata_limited,
            openpyxl_version=metadata.version("openpyxl"),
            defusedxml_available=importlib.util.find_spec("defusedxml") is not None,
            openpyxl_license=OPENPYXL_LICENSE,
            defusedxml_license=DEFUSEDXML_LICENSE,
            zip_uncompressed_bytes=zip_statistics.uncompressed_bytes,
            zip_compressed_bytes=zip_statistics.compressed_bytes,
            zip_compression_ratio=zip_statistics.compression_ratio,
        )

    def iter_sheet_rows(
        self,
        source_path: Path,
        sheet_name: str,
        inspection: SheetInspection,
    ) -> Iterator[SpreadsheetRow]:
        """data_only=Trueで保存済み計算結果をストリーミング取得する。"""
        workbook = _load_workbook(
            filename=source_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            worksheet = workbook[sheet_name]
            max_column = inspection.max_column or 1
            max_row = inspection.max_row or 1
            for row_number, row_values in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_column,
                    values_only=True,
                ),
                start=1,
            ):
                yield SpreadsheetRow(
                    sheet_name=sheet_name,
                    row_number=row_number,
                    cells=tuple(
                        SpreadsheetCell(
                            row_number=row_number,
                            column_number=column_number,
                            coordinate=f"{column_letter(column_number)}{row_number}",
                            value=value,
                        )
                        for column_number, value in enumerate(row_values, start=1)
                    ),
                )
        finally:
            workbook.close()


@dataclass(frozen=True, slots=True)
class _ZipStatistics:
    """OOXML ZIPコンテナの安全診断結果。"""

    uncompressed_bytes: int
    compressed_bytes: int
    compression_ratio: float


@dataclass(frozen=True, slots=True)
class _SheetMetadata:
    """通常モードで取得する詳細metadata。"""

    merged_ranges: tuple[str, ...]
    tables: tuple[XlsxTableInfo, ...]
    hidden_rows: tuple[int, ...]
    hidden_columns: tuple[int, ...]


def _inspect_ooxml_container(source_path: Path, options: XlsxParserOptions) -> _ZipStatistics:
    """展開せずZIPメタ情報だけで危険なXLSXを拒否する。"""
    if not source_path.is_file():
        raise XlsxWorkbookError("xlsx_not_ooxml", "通常ファイルではありません。")
    if source_path.stat().st_size <= 0:
        raise XlsxWorkbookError("xlsx_not_ooxml", "空ファイルです。")
    try:
        with zipfile.ZipFile(source_path) as archive:
            names = set(archive.namelist())
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise XlsxWorkbookError("xlsx_not_ooxml", "必須OOXMLメンバーがありません。")
            uncompressed_bytes = sum(member.file_size for member in archive.infolist())
            compressed_bytes = sum(member.compress_size for member in archive.infolist())
    except XlsxWorkbookError:
        raise
    except zipfile.BadZipFile as error:
        raise XlsxWorkbookError("xlsx_zip_unreadable", "ZIPとして読み取れません。") from error

    if uncompressed_bytes > options.max_uncompressed_bytes:
        raise XlsxWorkbookError(
            "xlsx_uncompressed_size_limit_exceeded",
            "ZIP展開サイズの安全上限を超えています。",
        )
    compression_ratio = uncompressed_bytes / max(compressed_bytes, 1)
    if compression_ratio > options.max_compression_ratio:
        raise XlsxWorkbookError(
            "xlsx_compression_ratio_limit_exceeded",
            "ZIP圧縮率の安全上限を超えています。",
        )
    return _ZipStatistics(
        uncompressed_bytes=uncompressed_bytes,
        compressed_bytes=compressed_bytes,
        compression_ratio=compression_ratio,
    )


def _collect_formulas_and_sheet_metrics(
    source_path: Path,
) -> tuple[dict[str, dict[str, str]], tuple[SheetInspection, ...]]:
    workbook = _load_workbook(
        filename=source_path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        formulas: dict[str, dict[str, str]] = {}
        sheets: list[SheetInspection] = []
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            sheet_formulas: dict[str, str] = {}
            non_empty_cell_count = 0
            min_row: int | None = None
            max_row: int | None = None
            min_column: int | None = None
            max_column: int | None = None
            declared_dimension = _safe_dimension(worksheet)
            row_limit = int(getattr(worksheet, "max_row", 1) or 1)
            column_limit = int(getattr(worksheet, "max_column", 1) or 1)
            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=row_limit,
                    min_col=1,
                    max_col=column_limit,
                    values_only=True,
                ),
                start=1,
            ):
                for column_number, value in enumerate(row, start=1):
                    if value is None or value == "":
                        continue
                    non_empty_cell_count += 1
                    min_row = row_number if min_row is None else min(min_row, row_number)
                    max_row = row_number if max_row is None else max(max_row, row_number)
                    min_column = (
                        column_number if min_column is None else min(min_column, column_number)
                    )
                    max_column = (
                        column_number if max_column is None else max(max_column, column_number)
                    )
                    if isinstance(value, str) and value.startswith("="):
                        sheet_formulas[f"{column_letter(column_number)}{row_number}"] = value
            formulas[worksheet.title] = dict(sorted(sheet_formulas.items()))
            sheets.append(
                SheetInspection(
                    sheet_name=worksheet.title,
                    sheet_index=sheet_index,
                    sheet_state=str(getattr(worksheet, "sheet_state", "visible")),
                    declared_dimension=declared_dimension,
                    actual_dimension=_actual_dimension(min_row, min_column, max_row, max_column),
                    min_row=min_row,
                    max_row=max_row,
                    min_column=min_column,
                    max_column=max_column,
                    non_empty_cell_count=non_empty_cell_count,
                    formula_cell_count=len(sheet_formulas),
                    merged_ranges=(),
                    tables=(),
                    hidden_rows=(),
                    hidden_columns=(),
                )
            )
        return formulas, tuple(sheets)
    finally:
        workbook.close()


def _collect_workbook_metadata(
    source_path: Path,
) -> tuple[dict[str, _SheetMetadata], tuple[DefinedNameInfo, ...]]:
    workbook = _load_workbook(
        filename=source_path,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        metadata_by_sheet: dict[str, _SheetMetadata] = {}
        for worksheet in workbook.worksheets:
            tables = tuple(
                sorted(
                    (_table_info(table) for table in worksheet.tables.values()),
                    key=lambda table: table.table_name,
                )
            )
            metadata_by_sheet[worksheet.title] = _SheetMetadata(
                merged_ranges=tuple(str(rng) for rng in worksheet.merged_cells.ranges),
                tables=tables,
                hidden_rows=tuple(
                    sorted(
                        row_index
                        for row_index, dimension in worksheet.row_dimensions.items()
                        if bool(getattr(dimension, "hidden", False))
                    )
                ),
                hidden_columns=tuple(
                    sorted(
                        column_index_from_string(column_letter_text)
                        for column_letter_text, dimension in worksheet.column_dimensions.items()
                        if bool(getattr(dimension, "hidden", False))
                    )
                ),
            )
        return metadata_by_sheet, _defined_names(workbook)
    finally:
        workbook.close()


def _load_workbook(**kwargs: object) -> Any:
    """openpyxl型をアダプター内に閉じ込め、mypy境界を1箇所にする。"""
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    return load_workbook(**kwargs)


def _merge_sheet_inspection(
    sheet: SheetInspection,
    metadata_value: _SheetMetadata | None,
) -> SheetInspection:
    if metadata_value is None:
        return sheet
    return SheetInspection(
        sheet_name=sheet.sheet_name,
        sheet_index=sheet.sheet_index,
        sheet_state=sheet.sheet_state,
        declared_dimension=sheet.declared_dimension,
        actual_dimension=sheet.actual_dimension,
        min_row=sheet.min_row,
        max_row=sheet.max_row,
        min_column=sheet.min_column,
        max_column=sheet.max_column,
        non_empty_cell_count=sheet.non_empty_cell_count,
        formula_cell_count=sheet.formula_cell_count,
        merged_ranges=metadata_value.merged_ranges,
        tables=metadata_value.tables,
        hidden_rows=metadata_value.hidden_rows,
        hidden_columns=metadata_value.hidden_columns,
    )


def _table_info(table: Any) -> XlsxTableInfo:
    table_range = str(getattr(table, "ref", ""))
    header_row = range_min_row(table_range)
    return XlsxTableInfo(
        table_name=str(getattr(table, "name", "")),
        display_name=str(getattr(table, "displayName", getattr(table, "name", ""))),
        range=table_range,
        header_row=header_row,
    )


def _defined_names(workbook: Any) -> tuple[DefinedNameInfo, ...]:
    values_method = getattr(workbook.defined_names, "values", None)
    defined_names = values_method() if callable(values_method) else ()
    records: list[DefinedNameInfo] = []
    for defined_name in defined_names:
        name = str(getattr(defined_name, "name", ""))
        destinations = getattr(defined_name, "destinations", ())
        try:
            for sheet_name, cell_range in destinations:
                records.append(
                    DefinedNameInfo(
                        defined_name=name,
                        destination_sheet=str(sheet_name),
                        destination_range=str(cell_range),
                    )
                )
        except (AttributeError, TypeError, ValueError):
            records.append(
                DefinedNameInfo(
                    defined_name=name,
                    destination_sheet=None,
                    destination_range=None,
                )
            )
    return tuple(
        sorted(records, key=lambda item: (item.defined_name, item.destination_sheet or ""))
    )


def _safe_dimension(worksheet: Any) -> str:
    calculate_dimension = getattr(worksheet, "calculate_dimension", None)
    if callable(calculate_dimension):
        try:
            return str(calculate_dimension(force=True))
        except TypeError:
            return str(calculate_dimension())
    return "A1:A1"


def _actual_dimension(
    min_row: int | None,
    min_column: int | None,
    max_row: int | None,
    max_column: int | None,
) -> str | None:
    if min_row is None or min_column is None or max_row is None or max_column is None:
        return None
    return f"{column_letter(min_column)}{min_row}:{column_letter(max_column)}{max_row}"


def column_letter(column_number: int) -> str:
    """1始まりの列番号をExcel列記号へ変換する。"""
    letters = ""
    value = column_number
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def column_index_from_string(column_letter_text: str) -> int:
    """Excel列記号を1始まりの列番号へ変換する。"""
    value = 0
    for character in column_letter_text.upper():
        if not ("A" <= character <= "Z"):
            break
        value = value * 26 + ord(character) - 64
    return value


def range_boundaries(range_text: str) -> tuple[int, int, int, int] | None:
    """A1:B2形式の範囲を列行境界へ変換する。"""
    match = re.fullmatch(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_text)
    if match is None:
        match = re.fullmatch(r"([A-Z]+)(\d+)", range_text)
        if match is None:
            return None
        column = column_index_from_string(match.group(1))
        row = int(match.group(2))
        return column, row, column, row
    return (
        column_index_from_string(match.group(1)),
        int(match.group(2)),
        column_index_from_string(match.group(3)),
        int(match.group(4)),
    )


def range_min_row(range_text: str) -> int | None:
    """範囲文字列から先頭行番号を返す。"""
    boundaries = range_boundaries(range_text)
    return None if boundaries is None else boundaries[1]


def _json_string_list(values: tuple[str, ...]) -> list[JsonValue]:
    return list(cast(list[JsonValue], list(values)))
