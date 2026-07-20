"""XLSX openpyxlアダプターの単体テスト。"""

import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from signate_drive_rag.ingestion.parsers.xlsx_adapter import (
    OpenpyxlWorkbookAdapter,
    XlsxParserOptions,
    XlsxWorkbookError,
)


def make_workbook(path: Path) -> None:
    """アダプターテスト用XLSXを作成する。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "スケジュール"
    sheet.append(["タスク", "日付"])
    sheet.append(["開始", "=DATE(2025,4,1)"])
    sheet.merge_cells("A4:B4")
    sheet["A4"] = "結合"
    sheet.row_dimensions[2].hidden = True
    sheet.column_dimensions["B"].hidden = True
    sheet.add_table(Table(displayName="ScheduleTable", ref="A1:B2"))
    hidden = workbook.create_sheet("非表示")
    hidden.sheet_state = "hidden"
    workbook.save(path)


def test_openpyxl_adapter_inspects_workbook_metadata(tmp_path: Path) -> None:
    """XLSXを読み、シート・数式・結合セル・テーブル・非表示情報を取得できる。"""
    path = tmp_path / "book.xlsx"
    make_workbook(path)

    inspection = OpenpyxlWorkbookAdapter().inspect(path, XlsxParserOptions())

    assert inspection.openpyxl_version == "3.1.5"
    assert inspection.defusedxml_available is True
    assert inspection.openpyxl_license == "MIT"
    assert inspection.defusedxml_license == "PSF"
    assert [sheet.sheet_name for sheet in inspection.sheets] == ["スケジュール", "非表示"]
    assert inspection.formulas["スケジュール"]["B2"] == "=DATE(2025,4,1)"
    assert inspection.sheets[0].merged_ranges == ("A4:B4",)
    assert inspection.sheets[0].tables[0].table_name == "ScheduleTable"
    assert inspection.sheets[0].hidden_rows == (2,)
    assert inspection.sheets[0].hidden_columns == (2,)
    assert inspection.sheets[1].sheet_state == "hidden"


def test_openpyxl_adapter_streams_cached_rows_without_modifying_source(tmp_path: Path) -> None:
    """data_only=Trueで行を読み、原本ファイルを変更しない。"""
    path = tmp_path / "book.xlsx"
    make_workbook(path)
    before = path.read_bytes()
    inspection = OpenpyxlWorkbookAdapter().inspect(path, XlsxParserOptions())

    rows = tuple(
        OpenpyxlWorkbookAdapter().iter_sheet_rows(path, "スケジュール", inspection.sheets[0])
    )

    assert rows[0].cells[0].coordinate == "A1"
    assert rows[1].cells[1].coordinate == "B2"
    assert path.read_bytes() == before


def test_openpyxl_adapter_rejects_empty_broken_and_non_ooxml_files(tmp_path: Path) -> None:
    """空ファイル、壊れたZIP、XLSXでないZIPを安全に拒否する。"""
    adapter = OpenpyxlWorkbookAdapter()
    empty = tmp_path / "empty.xlsx"
    empty.write_bytes(b"")
    broken = tmp_path / "broken.xlsx"
    broken.write_bytes(b"not zip")
    normal_zip = tmp_path / "normal.xlsx"
    with zipfile.ZipFile(normal_zip, "w") as archive:
        archive.writestr("hello.txt", "hello")

    with pytest.raises(XlsxWorkbookError, match="空ファイル"):
        adapter.inspect(empty, XlsxParserOptions())
    with pytest.raises(XlsxWorkbookError) as broken_error:
        adapter.inspect(broken, XlsxParserOptions())
    assert broken_error.value.issue_type == "xlsx_zip_unreadable"
    with pytest.raises(XlsxWorkbookError) as ooxml_error:
        adapter.inspect(normal_zip, XlsxParserOptions())
    assert ooxml_error.value.issue_type == "xlsx_not_ooxml"


def test_openpyxl_adapter_rejects_zip_safety_limits(tmp_path: Path) -> None:
    """展開サイズと異常圧縮率の上限超過を検出できる。"""
    path = tmp_path / "big.xlsx"
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
        archive.writestr("xl/workbook.xml", "<workbook/>")
        archive.writestr("xl/worksheets/sheet1.xml", "0" * 1000)

    with pytest.raises(XlsxWorkbookError) as size_error:
        OpenpyxlWorkbookAdapter().inspect(path, XlsxParserOptions(max_uncompressed_bytes=10))
    assert size_error.value.issue_type == "xlsx_uncompressed_size_limit_exceeded"

    with pytest.raises(XlsxWorkbookError) as ratio_error:
        OpenpyxlWorkbookAdapter().inspect(path, XlsxParserOptions(max_compression_ratio=1.0))
    assert ratio_error.value.issue_type == "xlsx_compression_ratio_limit_exceeded"
