"""XLSXパーサーの単体テスト。"""

from datetime import UTC, date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from signate_drive_rag.domain import SourceFile
from signate_drive_rag.ingestion.parsers.xlsx_parser import XlsxParser, XlsxParserError


def make_source_file(path: Path) -> SourceFile:
    """テスト用SourceFileを作成する。"""
    return SourceFile(
        path=path,
        relative_path=Path("日本語") / path.name,
        name=path.name,
        suffix=path.suffix.lower(),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size_bytes=path.stat().st_size,
        modified_at=datetime(2026, 7, 19, tzinfo=UTC),
    )


def make_complex_workbook(path: Path) -> None:
    """XLSXパーサーテスト用ワークブックを作成する。"""
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "スケジュール"
    schedule.append(["タスクID", "担当者", "開始日", "完了", "合計"])
    schedule.append(["T001", "佐藤", date(2025, 4, 1), True, "=SUM(1,2)"])
    schedule.append([])
    schedule.append(["項目", "値"])
    schedule.append(["費用", 4200000])
    schedule.merge_cells("A7:D7")
    schedule["A7"] = "結合セル"
    schedule.row_dimensions[2].hidden = True
    schedule.column_dimensions["D"].hidden = True
    schedule.add_table(Table(displayName="ScheduleTable", ref="A1:E2"))
    hidden = workbook.create_sheet("非表示")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "secret"
    workbook.create_sheet("空")
    workbook.defined_names.add(DefinedName("ScheduleArea", attr_text="'スケジュール'!$A$1:$E$2"))
    workbook.save(path)


def test_xlsx_parser_creates_workbook_sheet_and_row_block_units(tmp_path: Path) -> None:
    """XLSXを読み、要約unitと行ブロックunitを決定的に生成できる。"""
    path = tmp_path / "book.xlsx"
    make_complex_workbook(path)
    source_file = make_source_file(path)
    before = path.read_bytes()

    first = XlsxParser().parse(source_file)
    second = XlsxParser().parse(source_file)

    assert first.units == second.units
    assert path.read_bytes() == before
    assert first.parser_name == "openpyxl_xlsx"
    assert first.units[0].unit_type == "xlsx_workbook_summary"
    assert first.units[0].locator == "workbook"
    assert first.units[0].metadata["sheet_count"] == 3
    assert first.units[0].metadata["openpyxl_version"] == "3.1.5"
    assert first.units[0].metadata["defusedxml_available"] is True
    assert any(unit.unit_type == "xlsx_sheet_summary" for unit in first.units)
    row_units = [unit for unit in first.units if unit.unit_type == "xlsx_table_rows"]
    assert row_units
    assert row_units[0].locator.startswith("sheet:スケジュール/range:")
    assert row_units[0].metadata["sheet_name"] == "スケジュール"
    assert row_units[0].metadata["table_name"] == "ScheduleTable"
    assert "2025-04-01" in row_units[0].text
    assert "true" in row_units[0].text
    assert "formula: =SUM(1,2)" in row_units[0].text
    assert row_units[0].metadata["contains_hidden_rows"] is True
    assert row_units[0].metadata["contains_hidden_columns"] is True
    assert any(issue.issue_type == "xlsx_hidden_sheet" for issue in first.issues)
    assert any(issue.issue_type == "xlsx_sheet_has_no_cells" for issue in first.issues)
    assert any(issue.issue_type == "xlsx_formula_cached_value_missing" for issue in first.issues)


def test_xlsx_parser_supports_only_xlsx_case_insensitively(tmp_path: Path) -> None:
    """xlsxだけを大文字小文字に依存せず処理対象にする。"""
    path = tmp_path / "BOOK.XLSX"
    make_complex_workbook(path)
    source_file = make_source_file(path)

    assert XlsxParser().supports(source_file)
    assert not XlsxParser().supports(
        SourceFile(
            path=tmp_path / "book.xlsm",
            relative_path=Path("book.xlsm"),
            name="book.xlsm",
            suffix=".xlsm",
            mime_type=None,
            size_bytes=1,
            modified_at=datetime(2026, 7, 19, tzinfo=UTC),
        )
    )


def test_xlsx_parser_raises_safe_error_for_invalid_xlsx(tmp_path: Path) -> None:
    """不正なXLSXは安全な抽出失敗例外へ変換する。"""
    path = tmp_path / "broken.xlsx"
    path.write_text("not zip", encoding="utf-8")

    with pytest.raises(XlsxParserError, match="xlsx_zip_unreadable") as error:
        XlsxParser().parse(make_source_file(path))
    assert str(tmp_path) not in str(error.value)
